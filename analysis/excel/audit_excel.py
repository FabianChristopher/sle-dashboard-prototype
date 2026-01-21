#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def extract_unit(name: str | None) -> str | None:
    if not name:
        return None
    match = re.search(r"\(([^)]+)\)", name)
    return match.group(1).strip() if match else None


def clean_lab_name(name: str | None) -> str | None:
    if not name:
        return None
    return re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()


def iter_sheet_values(ws) -> list[tuple[Any, ...]]:
    return [row for row in ws.iter_rows(values_only=True)]


def find_row_index(rows: list[tuple[Any, ...]], predicate) -> int | None:
    for index_1_based, row in enumerate(rows, start=1):
        if predicate(row):
            return index_1_based
    return None


def top_left_preview(rows: list[tuple[Any, ...]], max_rows: int = 12, max_cols: int = 8) -> list[list[str | None]]:
    preview: list[list[str | None]] = []
    for row in rows[:max_rows]:
        preview.append([clean(c) for c in row[:max_cols]])
    return preview


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit the SLE prototype Excel workbook structure.")
    parser.add_argument(
        "--excel",
        default="renal bx pilot-mock patients.xlsx",
        help="Path to the Excel workbook (default: renal bx pilot-mock patients.xlsx)",
    )
    parser.add_argument(
        "--out",
        default="audit_excel_report.json",
        help="Path for JSON report output (default: audit_excel_report.json)",
    )
    parser.add_argument(
        "--sample",
        type=int,
        default=12,
        help="How many labs/domains to include as samples in stdout (default: 12)",
    )

    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True, read_only=True)

    report: dict[str, Any] = {
        "workbook": str(excel_path),
        "sheets": [],
    }

    print(f"Workbook: {excel_path.name}")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = iter_sheet_values(ws)

        col1 = [clean((row[0] if len(row) else None)) for row in rows]

        lab_header_row = None
        for i, v in enumerate(col1, start=1):
            if v == "Laboratory Findings":
                lab_header_row = i
                break

        domain_header_row = find_row_index(
            rows,
            lambda row: any(clean(cell) == "Y/N (active)" for cell in row),
        )

        sheet_entry: dict[str, Any] = {
            "name": sheet_name,
            "dims": {"rows": ws.max_row, "cols": ws.max_column},
            "lab_header_row": lab_header_row,
            "domain_header_row": domain_header_row,
            "timepoints": [],
            "labs": {"count": 0, "names": []},
            "domains": {"count": 0, "names": []},
            "preview_top_left": None,
        }

        tp_cols: list[int] = []
        tp_labels: list[str] = []

        if lab_header_row is not None:
            header = rows[lab_header_row - 1]
            header_vals = [clean(c) for c in header]
            for col_index_1_based, value in enumerate(header_vals, start=1):
                if col_index_1_based == 1:
                    continue
                if value is not None:
                    tp_cols.append(col_index_1_based)
                    tp_labels.append(value)

            sheet_entry["timepoints"] = tp_labels

            lab_names: list[str] = []
            r = lab_header_row + 1
            while r <= len(rows):
                row = rows[r - 1]
                name_raw = clean(row[0] if len(row) else None)
                if name_raw is None:
                    break
                lab_names.append(clean_lab_name(name_raw) or name_raw)
                r += 1

            unique_labs: list[str] = []
            seen: set[str] = set()
            for name in lab_names:
                if name and name not in seen:
                    unique_labs.append(name)
                    seen.add(name)

            sheet_entry["labs"] = {
                "count": len(unique_labs),
                "names": unique_labs,
            }

        if domain_header_row is not None and tp_cols:
            domain_names: list[str] = []
            r = domain_header_row + 1
            while r <= len(rows):
                row = rows[r - 1]
                names = [
                    clean(row[c - 1]) if (c - 1) < len(row) else None
                    for c in tp_cols
                ]
                domain_name = next((n for n in names if n is not None), None)
                if domain_name is None:
                    break
                domain_names.append(domain_name)
                r += 1

            unique_domains: list[str] = []
            seen: set[str] = set()
            for name in domain_names:
                if name and name not in seen:
                    unique_domains.append(name)
                    seen.add(name)

            sheet_entry["domains"] = {
                "count": len(unique_domains),
                "names": unique_domains,
            }

        if sheet_entry["lab_header_row"] is None and sheet_entry["domain_header_row"] is None:
            sheet_entry["preview_top_left"] = top_left_preview(rows)

        report["sheets"].append(sheet_entry)

        print(f"\n=== {sheet_name} ===")
        print(f"dims: {ws.max_row} rows x {ws.max_column} cols")
        print(f"lab_header_row: {lab_header_row if lab_header_row is not None else 'NOT FOUND'}")
        if tp_labels:
            print(f"timepoints: {tp_labels}")
        print(f"labs_count: {sheet_entry['labs']['count']}")
        if sheet_entry["labs"]["count"]:
            print(f"labs_sample: {sheet_entry['labs']['names'][: args.sample]}")
        print(f"domain_header_row: {domain_header_row if domain_header_row is not None else 'NOT FOUND'}")
        print(f"domains_count: {sheet_entry['domains']['count']}")
        if sheet_entry["domains"]["count"]:
            print(f"domains_sample: {sheet_entry['domains']['names'][: args.sample]}")

    out_path = Path(args.out).expanduser().resolve()
    out_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"\nWrote JSON report: {out_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
